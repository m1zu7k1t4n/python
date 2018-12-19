import openpyxl as px

wb = px.load_workbook('./2017_i4_seat.xlsx')
ws = wb['using']

with open("member_dict.py",'w',encoding='utf-8') as f:
    f.write("member_dict = {\n")
    for row in range(31,83):
            f.writelines(f'    "{ws.cell(row=row,column=1).value}" : ("{ws.cell(row=row,column=2).value}", "{ws.cell(row=row,column=3).value}"),\n')
    f.write("}\n")

